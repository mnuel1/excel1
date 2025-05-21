import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

def select_masterlist():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        masterlist_entry.delete(0, tk.END)
        masterlist_entry.insert(0, filepath)

def select_report():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        report_entry.delete(0, tk.END)
        report_entry.insert(0, filepath)

def generate_reports():
    masterlist_path = masterlist_entry.get()
    report_path = report_entry.get()

    if not os.path.exists(masterlist_path) or not os.path.exists(report_path):
        messagebox.showerror("Error", "Please select valid Excel files.")
        return

    try:
        process_reports(masterlist_path, report_path)
        messagebox.showinfo("Success", "All project reports have been generated.")
        os.startfile('project_reports')
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{e}")

def process_reports(masterlist_path, report_path):
    master_df = pd.read_excel(masterlist_path)
    report_df = pd.read_excel(report_path)

    report_df.columns = report_df.columns.str.strip().str.upper()
    report_df[['DATE', 'CODE', 'CUSTOMER']] = report_df[['DATE', 'CODE', 'CUSTOMER']].fillna(method='ffill')
    report_df = report_df.dropna(subset=['DOC. NO.'])
    report_df['CODE'] = report_df['CODE'].astype(str).str.strip()
    report_df['CUSTOMER'] = report_df['CUSTOMER'].astype(str).str.strip()

    output_dir = 'project_reports'
    os.makedirs(output_dir, exist_ok=True)

    projects = master_df['Project'].unique()

    for project in projects:
        customers = master_df[master_df['Project'] == project]['Name'].unique()
        filtered_report = report_df[
            (report_df['CODE'].astype(str).str.strip() + ' - ' + report_df['CUSTOMER'].astype(str).str.strip()).isin(customers)
        ]

        if filtered_report.empty:
            continue
        
        output_data = []
        report_df['UNIT PRICE'] = pd.to_numeric(report_df['UNIT PRICE'], errors='coerce').fillna(0)
        grand_total = 0

        for cust_code, group in filtered_report.groupby(filtered_report['CODE'] + ' - ' + filtered_report['CUSTOMER']):
            cust_total = group['UNIT PRICE'].sum()
            first_row = True

            for _, row in group.iterrows():
                line = {
                    '110204010104 - Other Receivables': cust_code if first_row else '',
                    'Balance': cust_total if first_row else '',
                    'Particulars': row['DOC. NO.'],
                    'Amount': row['UNIT PRICE'],
                }
                output_data.append(line)
                first_row = False

            output_data.append({
                '110204010104 - Other Receivables': '',
                'Balance': '',
                'Particulars': 'TOTAL',
                'Amount': cust_total,
            })

        grand_total = sum([row['Amount'] for row in output_data if row['Particulars'] == 'TOTAL'])
        output_data.append({
            '110204010104 - Other Receivables': 'TOTAL',
            'Balance': grand_total,
            'Particulars': '',
            'Amount': '',
        })        

        output_df = pd.DataFrame(output_data)
        output_file = os.path.join(output_dir, f'{project}_report.xlsx')
        output_df.to_excel(output_file, index=False)

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        default_font = Font(name='Arial Narrow', size=10, bold=False)
        bold_font = Font(name='Arial Narrow', size=10, bold=True)
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        thin_border = Border(bottom=Side(style='thin', color='000000'))
        alignment_left = Alignment(horizontal='left', vertical='center')
        alignment_right = Alignment(horizontal='right', vertical='center')
        
        header = [cell.value for cell in ws[1]]
        balance_col = header.index('Balance') + 1
        amount_col = header.index('Amount') + 1
            
        accounting_format = '_(#,##0.00_);_((#,##0.00);_("-"??_);_(@_)'
        
        for row in ws.iter_rows(min_row=2, min_col=balance_col, max_col=balance_col):
            for cell in row:
                cell.number_format = accounting_format

        for row in ws.iter_rows(min_row=2, min_col=amount_col, max_col=amount_col):
            for cell in row:
                cell.number_format = accounting_format

        wb.save(output_file)

# === GUI ===
root = tk.Tk()
root.title("Excel Report Generator")

tk.Label(root, text="Masterlist Excel File:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
masterlist_entry = tk.Entry(root, width=50)
masterlist_entry.grid(row=0, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=select_masterlist).grid(row=0, column=2, padx=10)

tk.Label(root, text="Report Excel File:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
report_entry = tk.Entry(root, width=50)
report_entry.grid(row=1, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=select_report).grid(row=1, column=2, padx=10)

tk.Button(root, text="Generate Reports", command=generate_reports, width=25, bg="green", fg="white").grid(row=2, column=0, columnspan=3, pady=20)

root.mainloop()
